# -*- coding: utf-8 -*-
import os
import re
import sys
from typing import Dict, List, Tuple
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from difflib import SequenceMatcher
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

# =========================
# 파일 경로 설정
# =========================
def get_file_paths():
    """명령행 인수에서 파일 경로를 가져오거나 사용법을 표시합니다."""
    if len(sys.argv) != 2:
        print("=" * 60)
        print("📋 엑셀 오류리스트 자동분류 프로그램")
        print("=" * 60)
        print()
        print("🔸 사용법:")
        print("  1. 엑셀 파일을 이 Python 파일 위에 드래그 앤 드롭하세요")
        print("  2. 또는 명령행에서: python error_list_auto_classify.py \"파일경로.xlsx\"")
        print()
        print("🔸 필요한 시트: 1차 점검, 2차 점검, 3차 점검, 품질점검")
        print("🔸 출력: 원본파일명_자동분류.xlsx")
        print()
        print("=" * 60)
        input("아무 키나 눌러 종료하세요...")
        sys.exit(1)
    
    src_path = sys.argv[1]
    
    # 파일 존재 확인
    if not os.path.exists(src_path):
        print(f"❌ 오류: 파일을 찾을 수 없습니다: {src_path}")
        input("아무 키나 눌러 종료하세요...")
        sys.exit(1)
    
    # 확장자 확인
    if not src_path.lower().endswith('.xlsx'):
        print(f"❌ 오류: Excel 파일(.xlsx)만 지원됩니다: {src_path}")
        input("아무 키나 눌러 종료하세요...")
        sys.exit(1)
    
    # 출력 파일 경로 생성
    base, ext = os.path.splitext(src_path)
    dst_path = f"{base}_자동분류{ext}"
    
    return src_path, dst_path

# 파일 경로 설정은 CLI 전용(main에서만 사용)

SHEETS = ["1차 점검", "2차 점검", "3차 점검", "품질점검"]
ALLOWED_TYPES_LATE = {"오탈자", "띄어쓰기"}   # 3차/품질 허용
ARROW = "→"
ARROWS = ["→","->","=>"]

DATA_START_ROW = 8
COL_A_TO_K_LAST = 11   # A~K
COL_ERR = 12           # L: 오류항목 (여러 줄)
COL_PAGE = 13          # M: 페이지번호 (여러 줄)
COL_FIX = 14           # N: 수정내역 (여러 줄, "이전 → 이후")

# 행 기준 종류별 최대 보존 개수, 그리고 행 전체 최대 보존 개수
MAX_PER_ERR_KIND = 5
MAX_PER_ROW_TOTAL = 5   # 최종 분배 대상은 행당 최대 5개

# 시트별 목표 분배 비율(전체 행 대비, "내용이 채워진 행 수" 기준).
# 2·3·4차는 이 비율을 쿼터(상한)로 사용하고, 1차는 상한 없는 catch-all이다.
#   - 1차: 항목이 있는 행이면 우선 채우고, 다른 시트 쿼터를 넘는 잉여 항목도 모두 흡수
#   - 2·3·4차: 목표 대비 가장 뒤처진 시트부터 한 항목씩 채워 비율에 수렴
#   - 3·4차는 ALLOWED_TYPES_LATE(오탈자/띄어쓰기) 항목만 받는 규칙 유지
SHEET_TARGET_RATIO = {1: 0.80, 2: 0.40, 3: 0.20, 4: 0.10}

# 중복 판단 기준 (유사도 임계값)
SIMILARITY_THRESHOLD = 0.85  # 85% 이상 유사하면 중복으로 판단

# 출력 폰트 (모든 시트 A~O열 8행~마지막행)
FONT_NAME = "맑은 고딕"
FONT_SIZE = 10

def ifont(color: str, bold: bool = False) -> InlineFont:
    """Rich Text 런용 InlineFont: 폰트·크기를 지정해 N열도 맑은 고딕 10pt로 보이게 한다."""
    return InlineFont(rFont=FONT_NAME, sz=FONT_SIZE, color=color, b=bold)

# 빨강 표시 시 bold도 함께 적용할 기호들 (.,:;''‘’“”" ·ㆍ)
RED_BOLD_PUNCT = set(".,:;'‘’\"“”·ㆍ")

def append_red(crt, text: str) -> None:
    """빨강 텍스트를 crt에 추가하되, 지정 기호(RED_BOLD_PUNCT)는 bold로도 처리한다.
    기호/비기호가 섞이면 같은 종류끼리 묶어 별도 런으로 나눈다."""
    if not text:
        return
    buf = ""
    buf_bold = False
    for ch in text:
        is_p = ch in RED_BOLD_PUNCT
        if buf and is_p != buf_bold:
            crt.append(TextBlock(text=buf, font=ifont("FF0000", bold=buf_bold)))
            buf = ""
        buf += ch
        buf_bold = is_p
    if buf:
        crt.append(TextBlock(text=buf, font=ifont("FF0000", bold=buf_bold)))

# =========================
# 유틸
# =========================
def normalize_error_type(v: str) -> str:
    v = (v or "").strip()
    if v == "링크오류" or v == "링크수정":
        return "링크"
    elif v == "단계오류" or  v == "단계수정":
        return "단계"
    elif v == "책갈피오류" or v == "석점줄임":
        return "책갈피수정"
    elif v=="추가" or v=="삭제" or v == "책갈피추가" or v == "책갈피삭제" or v == "책갈피추가·삭제":
        return "추가·삭제"
    else:
        return v

def split_lines(val) -> List[str]:
    if val is None:
        return []
    s = str(val)
    # Alt+Enter(\r\n), \n, \r 모두 인식해서 '줄 단위 항목'으로 분리
    parts = re.split(r'\r\n|\n|\r', s)
    # 앞뒤 공백만 정리하고, 빈 줄 제거
    return [p.strip() for p in parts if p.strip()]

def split_multi_fix(raw_fix: str) -> List[str]:
    """한 수정내역 텍스트 안에 '/' 기호가 있고 '→'(화살표)가 2개 이상이면, 그 셀에 수정내역이
    2개 이상 들어있는 것으로 보고 '/' 기준으로 분리해 여러 개로 돌려준다.
    (예: 'A → B / C → D' → ['A → B', 'C → D'])
    조건에 맞지 않으면 원본 그대로 1개만 돌려준다."""
    arrow_count = sum(raw_fix.count(a) for a in ARROWS)
    if '/' in raw_fix and arrow_count >= 2:
        return [p.strip() for p in raw_fix.split('/') if p.strip()]
    return [raw_fix]

def _is_sublist(small: List[str], big: List[str]) -> bool:
    """small이 big의 연속 부분 수열인지 여부."""
    n, m = len(small), len(big)
    if n == 0:
        return True
    if n > m:
        return False
    for i in range(m - n + 1):
        if big[i:i + n] == small:
            return True
    return False

def merge_fragments(fragments: List[str]) -> str:
    """'/'로 분리됐던 문장 조각들을 겹치는 단어를 찾아 하나의 문장으로 합친다.
    - 앞 조각의 '접미사 단어들'과 뒤 조각의 '접두사 단어들'이 겹치면 그 위치에서 합침
      (예: '〔그림 3-1〕' + '〔그림 3-1〕 연면적의' → '〔그림 3-1〕 연면적의')
    - 한쪽이 다른 쪽에 이미 포함되면 더 긴 쪽을 사용
    - 겹치는 단어가 없으면 공백으로 이어붙임 (예: '〔그림 3-1〕' + '연면적의' → '〔그림 3-1〕 연면적의')
    """
    parts = [f for f in fragments if f and f.strip()]
    if not parts:
        return ""
    merged = parts[0].split()
    for frag in parts[1:]:
        w = frag.split()
        if not w:
            continue
        if not merged:
            merged = w
            continue
        # merged의 접미사 == w의 접두사 인 최대 겹침 k 찾기
        overlap = 0
        for k in range(min(len(merged), len(w)), 0, -1):
            if merged[-k:] == w[:k]:
                overlap = k
                break
        if overlap:
            merged = merged + w[overlap:]
        elif _is_sublist(w, merged):
            pass  # w가 이미 merged에 포함됨
        elif _is_sublist(merged, w):
            merged = w  # merged가 w에 포함됨 → 더 긴 w 사용
        else:
            merged = merged + w  # 겹침 없음 → 공백으로 이어붙임
    return " ".join(merged)

def extract_page_numbers(page_str: str) -> List[str]:
    """페이지 문자열에서 모든 숫자 시퀀스 추출 (예: 'n.46' → ['46'])."""
    if not page_str:
        return []
    return re.findall(r'\d+', str(page_str))

def strip_matching_trailing_number(text: str, page_nums: List[str]) -> str:
    """text의 끝 번호가 page_nums 중 하나와 일치하면 그 번호와 앞쪽 공백을 제거."""
    if not text or not page_nums:
        return text
    m = re.search(r'\b(\d+)\s*$', text)
    if m and m.group(1) in page_nums:
        return text[:m.start()].rstrip()
    return text

def split_trailing_number(text: str) -> Tuple[str, str]:
    """끝 숫자를 (앞 텍스트, 끝 숫자)로 분리. 끝이 숫자가 아니면 (원문, '')."""
    m = re.search(r'^(.*?)\s*(\d+)\s*$', text or "")
    if m:
        return m.group(1).strip(), m.group(2)
    return (text or "").strip(), ""

# 번호체계 괄호 (ASCII + 전각): [] 〔〕 ［］ 【】 <> 〈〉 《》 () （）
_BR_OPEN  = r'[\[〔［【<〈《(（]'
_BR_CLOSE = r'[\]〕］】>〉》)）]'
_BR_INNER = r'[^\[\]〔〕［］【】<>〈〉《》()（）]'
_BRACKET_RE = re.compile(r'^\s*' + _BR_OPEN + r'(' + _BR_INNER + r'*)' + _BR_CLOSE + r'(.*)$', re.S)

def is_bracket_to_dot(before: str, after: str) -> bool:
    """선두 괄호 '[...]'를 '....'(온점)으로 바꾼 변환인지 판정 (한 방향).
    괄호는 [] <> () 및 전각 〔〕 ［］ 【】 〈〉 《》 （） 를 모두 인식한다.
    예: '〔그림 1-1〕 硏究의' → '그림 1-1. 硏究의' → True
    """
    if not before or not after:
        return False
    m = _BRACKET_RE.match(before)
    if not m:
        return False
    converted = f"{m.group(1)}.{m.group(2)}"
    return " ".join(converted.split()) == " ".join(after.split())

def is_bracket_dot_change(before: str, after: str) -> bool:
    """번호체계 괄호↔온점 변환인지 (양방향).
    - 정방향 '[...]' → '....'  (before가 괄호형)
    - 역방향 '....' → '[...]'  (after가 괄호형)
    """
    return is_bracket_to_dot(before, after) or is_bracket_to_dot(after, before)

# '단계' 수정내역 접두사: '[뎁스 X→Y] TEXT' (괄호·공백 가변)
DAN_PREFIX_RE = re.compile(r'^\s*\[\s*뎁스\s*(\d+)\s*(?:→|->|=>)\s*(\d+)\s*\]\s*(.*)$', re.S)

def parse_dan(raw_fix: str):
    """'단계' 수정내역 '[뎁스 X→Y] TEXT'를 (from, to, text)로 파싱. 형식이 안 맞으면 None."""
    m = DAN_PREFIX_RE.match(raw_fix or "")
    if not m:
        return None
    return m.group(1), m.group(2), m.group(3).strip()

def parse_fix(raw_fix: str, err_type: str, page) -> Tuple[str, str, object]:
    """수정내역 문자열 하나를 (before, after, arrow)로 파싱한다.
    - 화살표(→)로 before/after 분리 (없으면 before='', after=전체)
    - '링크'가 아니면 좌/우 끝 번호가 페이지번호와 같을 때 제거
    - '추가·삭제' 타입은 after 첫 문자 '-'/'+'를 '삭제 '/'추가 '로 변환
    - '단계' 타입은 '[뎁스 X→Y] TEXT'에서 TEXT만 추출 (괄호 안 화살표로 잘못 분리되지 않게)
    """
    if err_type == "단계":
        parsed = parse_dan(raw_fix)
        if parsed:
            _from, _to, text = parsed
            return text, text, None  # 중복판정·표시용 본문 텍스트

    found_arrow = next((a for a in ARROWS if a in raw_fix), None)
    if found_arrow:
        left, right = raw_fix.split(found_arrow, 1)
        before, after = left.strip(), right.strip()
    else:
        before, after = "", str(raw_fix).strip()

    if err_type != "링크":
        page_nums = extract_page_numbers(page)
        if page_nums:
            before = strip_matching_trailing_number(before, page_nums)
            after = strip_matching_trailing_number(after, page_nums)

    if err_type == "추가·삭제" and after:
        if after[0] == '-':
            after = "삭제 " + after[1:].lstrip()
        elif after[0] == '+':
            after = "추가 " + after[1:].lstrip()

    return before, after, found_arrow

def merged_before_after(raw_fix: str, err_type: str, page) -> Tuple[str, str, object]:
    """raw_fix를 '/' 기준으로 분리·파싱한 뒤, 여러 조각이면 겹치는 단어로 병합해
    '깨끗한' (before, after, arrow)를 돌려준다. 중복판정·렌더링 모두 이 결과를 쓴다.
    (예: '[그림 1-1] → 그림 1-1. / [그림 1-1] 연구설계 → 그림 1-1. 연구설계'
        → before '[그림 1-1] 연구설계', after '그림 1-1. 연구설계')
    """
    subs = []
    for s in split_multi_fix(raw_fix):
        b, a, arrow = parse_fix(s, err_type, page)
        if b or a:
            subs.append((b, a, arrow))
    if len(subs) > 1:
        mb = merge_fragments([x[0] for x in subs])
        ma = merge_fragments([x[1] for x in subs])
        arrow = next((x[2] for x in subs if x[2]), None)
        return mb, ma, arrow
    if subs:
        return subs[0]
    return parse_fix(raw_fix, err_type, page)

def calculate_similarity(item1: Dict, item2: Dict) -> float:
    """두 오류 항목 간의 유사도를 계산합니다."""
    # 오류 유형이 다르면 중복이 아님
    if item1["err"] != item2["err"]:
        return 0.0

    b1, b2 = item1["before"], item2["before"]
    a1, a2 = item1["after"], item2["after"]
    after_sim = SequenceMatcher(None, a1, a2).ratio()

    # 추가·삭제처럼 before가 양쪽 모두 비면 before_sim이 항상 1.0이 되어 점수가 부풀려진다
    # (공통 접두 '삭제 (1) ' 때문에 다른 내용도 85%를 넘김). 이 경우 after만으로 판단.
    if not b1 and not b2:
        return after_sim

    before_sim = SequenceMatcher(None, b1, b2).ratio()
    # 수정 전후 내용의 평균 유사도 반환
    return (before_sim + after_sim) / 2.0

def change_signature(item: Dict):
    """항목의 '변경 패턴' 서명을 반환한다. 문맥은 보지 않고, 오직 '무엇이 어떻게 바뀌었는가'
    (바뀐 글자 자체)와 오류유형만으로 서명을 만든다. 숫자는 무시(정규화)한다.
    같은 서명이면 '같은 패턴의 반복 수정'으로 보고 하나만 남긴다.
    예) '대퇴위·좌의 변화 → 대퇴위ㆍ좌의 변화', '신근·좌(Nm)의 변화 → 신근ㆍ좌(Nm)의 변화'
        → 둘 다 '·→ㆍ' 변경뿐이므로 (문맥이 달라도) 동일 서명.
    변경이 없거나(둘이 같음) 숫자 차이뿐이면 None(패턴 dedup 대상 아님).
    """
    err = item.get("err", "")
    before = item.get("before") or ""
    after = item.get("after") or ""
    if before == after:
        return None
    nb = re.sub(r'\d+', '#', before)   # 숫자 런을 '#'로 정규화
    na = re.sub(r'\d+', '#', after)
    if nb == na:
        return None  # 숫자만 다른 경우는 패턴 dedup 대상에서 제외

    sig = [err]
    has_change = False
    for tag, i1, i2, j1, j2 in SequenceMatcher(None, nb, na).get_opcodes():
        if tag == "equal":
            continue
        has_change = True
        sig.append((tag, nb[i1:i2], na[j1:j2]))  # 문맥 없이 바뀐 부분만
    return tuple(sig) if has_change else None

# 순서 표지(목록 번호) 인식용
_CIRCLED_NUM = "①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳"   # 원문자 ①=1
_HANGUL_SYL  = "가나다라마바사아자차카타파하"               # 가=1, 나=2 …
_HANGUL_JAMO = "ㄱㄴㄷㄹㅁㅂㅅㅇㅈㅊㅋㅌㅍㅎ"               # ㄱ=1, ㄴ=2 …

def list_marker(text: str):
    """문자열 앞부분의 '목록 순서표지'를 (그룹키, 순서값)으로 반환. 없으면 (None, None).
    표지 앞에 짧은 라벨(부록/표/그림/삭제/( 등)이 와도 인식한다.
    그룹키 = (라벨, 표지종류) — 오류유형·뒤 내용과 무관하게 같은 라벨·형식이면 한 목록으로 본다.
    인식: 라벨(1) / 라벨1) / 라벨1. / ① / 가) / 가. / ㄱ) / ㄱ.
    """
    s = text or ""
    m = re.match(r'\s*(\D{0,6}?)\((\d+)\)(?=\s|$)', s)            # 라벨(1)
    if m:
        return (m.group(1).strip(), "()"), int(m.group(2))
    m = re.match(r'\s*(\D{0,6}?)(\d+)\s*([)\.])(?=\s|$)', s)      # 라벨1)  라벨1.
    if m:
        return (m.group(1).strip(), m.group(3)), int(m.group(2))
    m = re.match(r'\s*(\D{0,6}?)([①-⑳])', s)                     # 라벨①
    if m:
        return (m.group(1).strip(), "①"), _CIRCLED_NUM.index(m.group(2)) + 1
    m = re.match(r'\s*([가-힣])\s*([)\.])(?=\s|$)', s)            # 가)  가.
    if m and m.group(1) in _HANGUL_SYL:
        return ("", "가" + m.group(2)), _HANGUL_SYL.index(m.group(1)) + 1
    m = re.match(r'\s*([ㄱ-ㅎ])\s*([)\.])(?=\s|$)', s)            # ㄱ)  ㄱ.
    if m and m.group(1) in _HANGUL_JAMO:
        return ("", "ㄱ" + m.group(2)), _HANGUL_JAMO.index(m.group(1)) + 1
    return None, None

def _char_class(ch: str) -> str:
    """글자의 종류: H=한글 C=한자 L=라틴 N=숫자(#) S=공백 P=기타기호."""
    if ch.isspace():
        return "S"
    if ("가" <= ch <= "힣") or ("ㄱ" <= ch <= "ㅣ"):
        return "H"
    if ("一" <= ch <= "鿿") or ("㐀" <= ch <= "䶿") or ("豈" <= ch <= "﫿"):
        return "C"
    if ch == "#" or ch.isdigit():
        return "N"
    if "a" <= ch.lower() <= "z":
        return "L"
    return "P"

def _classify(s: str):
    """문자열에 든 글자종류의 '집합'(공백 제외). 단어 수·길이·순서는 무시.
    예: '사과'→('H',), '砂果'→('C',), '예비 영어 설문지'→('H',), '사전 영어 쓰기 시험'→('H',)."""
    return tuple(sorted({_char_class(ch) for ch in s if not ch.isspace()}))

def _abstract_diff(item: Dict):
    """change_signature의 바뀐 글자들을 '글자종류'로 추상화한 서명. 같은 종류의 변경이면 동일.
    예: 삭제 '예비 영어 설문지'/'사전 영어 쓰기 시험' → 둘 다 (delete, 한글, '') 로 동일."""
    cs = change_signature(item)
    if cs is None:
        return None
    return (cs[0],) + tuple((tag, _classify(b), _classify(a)) for tag, b, a in cs[1:])

def conversion_signature(item: Dict):
    """변경이 '교차-스크립트 변환'(한글↔한자↔영문처럼 글자 종류 자체가 바뀌는 replace)만으로
    이루어지면, 글자종류로 추상화한 서명을 반환(번호와 무관하게 같은 변환이면 동일). 아니면 None.
    예: '사과→砂果','배→拜','대한→大韓' → 모두 한글→한자 변환 → 동일 서명.
        '가→나'(한글→한글)·'·→ㆍ'(기호) 등은 None(정확 일치로만 판정).
    """
    cs = change_signature(item)
    if cs is None or len(cs) < 2:
        return None
    abst = []
    for tag, b, a in cs[1:]:
        if tag != "replace":
            return None  # 삽입/삭제가 섞이면 순수 변환이 아님
        cb = {c for c in map(_char_class, b) if c in "HCL"}   # 언어 스크립트만(한글/한자/라틴)
        ca = {c for c in map(_char_class, a) if c in "HCL"}
        if not cb or not ca or cb == ca:
            return None  # 한쪽이 언어가 아니거나(기호 등) 같은 스크립트면 변환 아님
        abst.append((tuple(sorted(cb)), tuple(sorted(ca))))
    return (cs[0], tuple(abst))

def _list_number_key(item: Dict):
    """연속-목록 dedup용 (그룹키, 순서값) 또는 (None, None).
    '같은 오류유형 + 같은 diff 처리 + 연속 번호'를 중복으로 본다.

    텍스트가 바뀌는 유형(오탈자·띄어쓰기·책갈피수정)은 change_signature(=바뀐 부분)로 이미
    내용·번호 무관하게 중복 판정되므로 여기서 다루지 않는다. diff가 텍스트 편집이 아닌
    두 유형만 처리한다:
      - 추가·삭제: 내용 자체가 삽입/삭제 → 내용 무시, '추가/삭제' 처리 + 목록표지 번호로 판정.
      - 단계: 본문 불변(변경=단계 전환) → 본문(숫자정규화)+단계전환이 같고 번호가 연속이면 중복.
    """
    err = item.get("err", "")
    before = item.get("before") or ""
    after = item.get("after") or ""

    # 단계: 본문 불변(변경=단계 전환). 본문 숫자로 순서, (본문 정규화 + 단계 전환)으로 그룹.
    if err == "단계" and before and before == after:
        m = re.search(r'\d+', before)
        if not m:
            return None, None
        dan = parse_dan(item.get("raw_fix") or "")
        dsig = (dan[0], dan[1]) if dan else ()
        return (err, "단계", re.sub(r'\d+', '#', before), dsig), int(m.group())

    # 추가·삭제: 내용 전체가 삽입/삭제 → 내용 무시, '삭제/추가' 처리 + 목록번호로 판정.
    if err == "추가·삭제":
        _k, seq = list_marker(after)
        if seq is None:
            return None, None
        pre = "삭제" if after.startswith("삭제") else ("추가" if after.startswith("추가") else "")
        return (err, "addel", pre), seq

    # 그 외: diff에 'replace'(특정 글자 교체)가 있으면 그 자체가 식별자 → 정확 비교(change_signature)
    # 또는 변환 비교(conversion_signature)에 맡긴다. replace가 없으면(삽입/삭제형) 삽입·삭제 내용이
    # 가변 목록 항목일 수 있으므로, 목록 번호가 연속이면 같은 처리의 반복으로 보고 중복.
    cs = change_signature(item)
    has_replace = cs is not None and any(op[0] == "replace" for op in cs[1:])
    if not has_replace:
        mk, seq = list_marker(after)
        if seq is None and before:
            mk, seq = list_marker(before)
        if seq is not None:
            return (mk, _abstract_diff(item)), seq

    return None, None

def dedup_consecutive_numbered(items: List[Dict]) -> List[Dict]:
    """목록 순서표지 항목이 '연속'이면 첫 표지만 남기고 나머지는 제거한다.
    순서값이 직전보다 작거나 같으면(=다시 1부터 시작 등) 새 그룹으로 보고 유지한다.
    예) (1)동북 (2)서북 (3)서남 (4)중남부 → (1)동북만 남김
        (1)동북 (2)서북 / (1)추가내용 → (1)동북, (1)추가내용 유지
    """
    result = []
    last_num: Dict = {}   # 그룹키 -> 직전 번호
    for it in items:
        key, num = _list_number_key(it)
        if key is None:
            result.append(it)
            continue
        prev = last_num.get(key)
        if prev is not None and num > prev:
            last_num[key] = num   # 연속 번호 → 제거(유지 목록에 안 넣음)
            continue
        result.append(it)         # 첫 항목 또는 번호 리셋 → 유지
        last_num[key] = num
    return result

def remove_duplicates(items: List[Dict]) -> List[Dict]:
    """중복된 오류 항목을 제거하고 고유한 항목들만 반환합니다.
    (0) 번호체계 '(N)' 연속 항목은 첫 번호만 남긴다(리스트형 반복 삭제/추가).
    (1) 변경 패턴 서명이 같으면(번호만 다른 동일 수정의 반복) 하나만 남긴다.
    (2) 그 외에는 기존처럼 수정전/후 전체 텍스트 85% 이상 유사면 중복으로 본다.
    """
    if not items:
        return items

    items = dedup_consecutive_numbered(items)   # (0) 번호체계 연속 정리

    unique_items = []
    seen_sigs = set()
    seen_conv = set()

    for current_item in items:
        # (1) 변경 패턴이 완전히 같으면 중복 (번호 무관)
        sig = change_signature(current_item)
        if sig is not None and sig in seen_sigs:
            continue

        # (1') 교차-스크립트 변환(한글↔한자 등)은 같은 변환이면 중복 (번호 무관)
        conv = conversion_signature(current_item)
        if conv is not None and conv in seen_conv:
            continue

        # (2) 전체 텍스트 유사도(85%) 기준 중복
        is_duplicate = any(
            calculate_similarity(current_item, ex) >= SIMILARITY_THRESHOLD
            for ex in unique_items
        )
        if is_duplicate:
            continue

        unique_items.append(current_item)
        if sig is not None:
            seen_sigs.add(sig)
        if conv is not None:
            seen_conv.add(conv)

    return unique_items

def append_diff_blocks(crt, before: str, after: str, side: str = "after") -> None:
    """before↔after를 비교해 side('after' 또는 'before')에 해당하는 문자열을 색칠해 crt에 추가.
    같은 부분은 검정, 바뀐 부분은 빨강 TextBlock으로 추가한다.

    - 변경된 글자는 '그 글자가 실제로 존재하는 쪽'에서만 빨강으로 표시한다.
      (after면 추가/수정 글자, before면 삭제/수정 글자)
    - 반대쪽에서만 일어난 변경(그 자리에 side쪽 글자가 없음)은 보통 강조하지 않는다.
      단, 그 변경 내용이 '공백만'(띄어쓰기 추가/삭제)이면 양쪽 모두 위치를 보여야 하므로
      side쪽에서도 변경 경계의 앞뒤 글자를 빨강으로 표시한다.
      (예: 콜론 삭제는 before에만, 공백 삭제/추가는 before·after 모두)

    바뀐 부분이 '공백만'인 경우(예: 띄어쓰기), 그 공백은 단독 런 `<t> </t>`이 되어
    xml:space="preserve"가 빠진 채 저장되고 Excel이 파일을 열 때 잘라내 버린다.
    이를 막기 위해 공백만 든 변경 런은 앞뒤 한 글자씩을 같은 빨강 런으로 흡수해
    공백이 항상 런 '내부'(글자 사이)에 오도록 한다. 흡수된 앞뒤 글자도 빨강으로 강조된다.
    """
    if side == "before":
        text = before
        own_tags = ("delete", "replace")   # before에 글자가 있는 변경
        gap_tag = "insert"                 # before엔 글자가 없는 변경(반대쪽에서만 추가)
    else:
        text = after
        own_tags = ("insert", "replace")   # after에 글자가 있는 변경
        gap_tag = "delete"                 # after엔 글자가 없는 변경(반대쪽에서만 삭제)

    flags = [False] * len(text)
    gap_points: List[int] = []  # '공백만' 변경이 반대쪽에서 일어난 경계 인덱스
    for tag, i1, i2, j1, j2 in SequenceMatcher(None, before, after).get_opcodes():
        if tag in own_tags:
            # 이 side에 실제 존재하는 변경 글자 → 빨강
            lo, hi = (i1, i2) if side == "before" else (j1, j2)
            for k in range(lo, hi):
                flags[k] = True
        elif tag == gap_tag:
            # 반대쪽에서만 일어난 변경. 그 내용이 '공백만'(띄어쓰기)일 때만 이 side에도 위치 표시.
            gap_content = after[j1:j2] if side == "before" else before[i1:i2]
            if gap_content != "" and gap_content.strip() == "":
                gap_points.append(i1 if side == "before" else j1)

    # 공백(띄어쓰기) 변경 위치의 앞뒤 글자를 빨강으로 (문자열 내부 변경만)
    for p in gap_points:
        if 0 < p < len(text):
            flags[p] = True          # 경계 오른쪽 글자
            flags[p - 1] = True      # 경계 왼쪽 글자

    # 연속 동일 플래그를 런으로 묶기
    segs: List[List] = []  # [text, is_changed]
    for ch, fl in zip(text, flags):
        if segs and segs[-1][1] == fl:
            segs[-1][0] += ch
        else:
            segs.append([ch, fl])

    # 공백만 든 변경 런이 단독으로 남지 않도록 앞뒤 런에서 1글자씩 흡수
    for k, cur in enumerate(segs):
        if not cur[1] or cur[0].strip() != "":
            continue
        # 앞 런의 마지막 글자를 빨강 런 앞쪽으로
        if k > 0 and segs[k - 1][0]:
            cur[0] = segs[k - 1][0][-1] + cur[0]
            segs[k - 1][0] = segs[k - 1][0][:-1]
        # 뒤 런의 첫 글자를 빨강 런 뒤쪽으로
        if k < len(segs) - 1 and segs[k + 1][0]:
            cur[0] = cur[0] + segs[k + 1][0][0]
            segs[k + 1][0] = segs[k + 1][0][1:]

    for text, changed in segs:
        if not text:
            continue
        if changed:
            append_red(crt, text)  # 빨강 + 지정 기호 bold
        else:
            crt.append(TextBlock(text=text, font=ifont("000000")))


def render_fix_blocks(crt, u: Dict) -> None:
    """수정내역 단위 하나(u)를 crt에 렌더링한다.
    - 일반: 수정전 → 수정후 양쪽에 바뀐 글자 빨강 diff
    - '추가·삭제': diff 없이 '추가'/'삭제' 접두사만 빨강, 나머지 검정
    - '단계': 'TEXT'(검정) 뒤에 'X단계 → Y단계' 접미사(단계 부분만 빨강) 부착
    - '링크': before/after 끝 숫자가 페이지번호 → '{본문} {before번호} → {after번호}'
      (페이지번호 부분만 빨강)
    """
    err_type = u.get("err", "")
    before = u.get("before") or ""
    after = u.get("after") or ""
    arrow = u.get("arrow")

    # '단계': 본문(검정) + 접미사 'X단계 → Y단계' (단계 부분만 빨강)
    if err_type == "단계" and u.get("dan_from") is not None:
        text = before  # = 본문 TEXT
        if text:
            crt.append(TextBlock(text=text + " ", font=ifont("000000")))
        append_red(crt, f"{u.get('dan_from')}단계")
        crt.append(TextBlock(text=" → ", font=ifont("808080")))
        append_red(crt, f"{u.get('dan_to')}단계")
        return

    # '링크': '{본문} {before페이지번호} → {after페이지번호}' (페이지번호만 빨강)
    if err_type == "링크":
        btext, bnum = split_trailing_number(before)
        _atext, anum = split_trailing_number(after)
        if bnum and anum:
            if btext:
                crt.append(TextBlock(text=btext + " ", font=ifont("000000")))
            append_red(crt, f"n.{bnum}")
            crt.append(TextBlock(text=" → ", font=ifont("808080")))
            append_red(crt, anum)
            return
        # 끝 숫자가 없으면 일반 diff로 폴백

    is_add_del = (err_type == "추가·삭제")

    # 추가·삭제 항목에 접두사가 없으면 기본 "추가 " 추가 ('삭제 '는 그대로 유지)
    if is_add_del:
        if not (after.startswith("추가 ") or after.startswith("삭제 ")):
            after = f"추가 {after}"

    arrow_block = f" {arrow or ARROW} "

    # 추가·삭제 접두사('추가 '/'삭제 ')는 주석성 표기이므로 before 비교는 접두사 제외
    ad_prefix = None
    if is_add_del:
        if after.startswith("추가 "):
            ad_prefix = "추가 "
        elif after.startswith("삭제 "):
            ad_prefix = "삭제 "
    compare_after = after[len(ad_prefix):] if ad_prefix else after

    # 수정전(before): 추가·삭제는 검정 통짜, 그 외는 삭제·수정 글자를 빨강 diff
    if before:
        if is_add_del or before == compare_after:
            crt.append(TextBlock(text=before, font=ifont("000000")))
        else:
            append_diff_blocks(crt, before, compare_after, side="before")
        crt.append(TextBlock(text=arrow_block, font=ifont("808080")))

    # 수정후(after): '추가 '/'삭제 ' 접두사만 빨강
    if ad_prefix:
        append_red(crt, ad_prefix)
    # 추가·삭제는 나머지 내용을 검정 통짜, 그 외는 추가·수정 글자를 빨강 diff
    if is_add_del or before == compare_after:
        crt.append(TextBlock(text=compare_after, font=ifont("000000")))
    else:
        append_diff_blocks(crt, before, compare_after, side="after")


def expand_fix_units(items: List[Dict]) -> List[Dict]:
    """분배된 항목들을 '/' 기준으로 펼쳐 수정내역 단위 리스트로 만든다.
    각 단위: {err, page, before, after, arrow}. 한 셀의 L/M/N 열을 같은 줄 수로 맞추기 위해
    오류항목·페이지 열도 이 결과를 기준으로 기록한다.
    """
    units: List[Dict] = []
    for it in items:
        err_type = it.get("err", "")
        page = it.get("page", "")
        raw_fix = it.get("raw_fix")

        # 이 항목에서 나온 수정내역 단위들 수집 (각 sub는 before/after/arrow[/dan_from,dan_to] 딕셔너리)
        subs: List[Dict] = []
        if raw_fix is None:
            before, after, arrow = it.get("before") or "", it.get("after") or "", it.get("arrow")
            if before or after:
                subs.append({"before": before, "after": after, "arrow": arrow})
        elif err_type == "단계":
            # '단계'는 괄호 안 화살표 때문에 '/' 분리를 하지 않고 통째로 파싱
            parsed = parse_dan(raw_fix)
            if parsed:
                df, dt, text = parsed
                subs.append({"before": text, "after": text, "arrow": None, "dan_from": df, "dan_to": dt})
            else:
                before, after, arrow = parse_fix(raw_fix, err_type, page)
                if before or after:
                    subs.append({"before": before, "after": after, "arrow": arrow})
        else:
            # '/'로 분리된 조각들은 하나의 문장 → 겹치는 단어로 병합 (중복판정과 동일 기준)
            before, after, arrow = merged_before_after(raw_fix, err_type, page)
            if before or after:
                subs.append({"before": before, "after": after, "arrow": arrow})

        # 항목의 첫 단위만 item_start=True (L/M 열에 값을 한 번만 넣기 위함).
        # '/'로 분리된 같은 항목의 이어지는 단위는 item_start=False → L/M 빈칸.
        for j, s in enumerate(subs):
            units.append({"err": err_type, "page": page, "item_start": (j == 0), **s})
    return units


def apply_rich_diff_for_bidir(cell, units):
    """이미 펼쳐진 수정내역 단위 리스트(units)를 줄마다 빨강 diff로 렌더링한다."""
    if not units:
        cell.value = None
        return

    crt = CellRichText()
    for u in units:
        line_blocks: List = []
        render_fix_blocks(line_blocks, u)
        if not line_blocks:
            continue
        # 줄바꿈은 단독 런(<t>\n</t>)이면 openpyxl이 preserve를 안 붙여 Excel이 잘라낸다.
        # 따라서 새 줄의 첫 블록 텍스트 앞에 '\n'을 붙여(비공백 런에 실어) 보존되게 한다.
        if crt:
            first = line_blocks[0]
            if isinstance(first, TextBlock):
                line_blocks[0] = TextBlock(text="\n" + str(first.text), font=first.font)
            else:
                line_blocks[0] = "\n" + str(first)
        crt.extend(line_blocks)

    cell.value = crt
    cell.alignment = Alignment(wrap_text=True, vertical='top')





def compute_last_row(ws1) -> int:
    last = DATA_START_ROW - 1
    for r in range(DATA_START_ROW, ws1.max_row + 1):
        # 제어번호(B열)나 L/M/N 중 하나라도 값이 있으면 데이터 행으로 인정
        if any(ws1.cell(r, c).value not in (None, "") for c in (2, COL_ERR, COL_PAGE, COL_FIX)):
            last = r
    return last

def merge_inputs_from_ws1_ws2_ws4(ws1, ws2, ws4) -> List[Dict]:
    """1차 점검(ws1)·2차 점검(ws2)·품질점검(ws4)의 raw 입력을 제어번호(B열) 기준으로
    메모리상에서 병합하여 리스트로 반환한다. 시트에는 아무것도 쓰지 않는다.

    반환: [{"a_to_k": [v1..v11], "err": [str,...], "page": [str,...], "fix": [str,...]}, ...]

    - ws1 등장 순서를 앞에 두고, 이후 ws2-only, ws4-only 제어번호를 차례로 이어붙임
    - A~K열은 세 시트 모두 동일 내용이라는 전제 하에 첫 등장 행의 값을 캡처
    - L/M/N은 세 시트의 모든 행에서 split_lines로 분리해 누적
    - 같은 제어번호가 비인접/여러 행에 분산되어도 모두 한 그룹으로 합쳐짐
    - 제어번호가 비어 있는 행은 절대 다른 행과 병합되지 않음(고유키 사용)
    - 완전 빈 행(B/L/M/N 모두 비어있음)은 무시
    """
    last_row_1 = compute_last_row(ws1)
    last_row_2 = compute_last_row(ws2)
    last_row_4 = compute_last_row(ws4)

    groups: Dict[object, Dict] = {}

    def absorb(ws, last_row: int, sheet_id: int) -> None:
        for r in range(DATA_START_ROW, last_row + 1):
            raw_b = ws.cell(r, 2).value
            ctrl_str = "" if raw_b is None else str(raw_b).strip()
            err_lines  = split_lines(ws.cell(r, COL_ERR).value)
            page_lines = split_lines(ws.cell(r, COL_PAGE).value)
            fix_lines  = split_lines(ws.cell(r, COL_FIX).value)

            # 완전 빈 행(제어번호도 없고 L/M/N도 비어있음)은 무시
            if not ctrl_str and not err_lines and not page_lines and not fix_lines:
                continue

            key = ctrl_str if ctrl_str else ("__nokey__", sheet_id, r)

            if key not in groups:
                groups[key] = {
                    "a_to_k": [ws.cell(r, c).value for c in range(1, COL_A_TO_K_LAST + 1)],
                    "err":  [],
                    "page": [],
                    "fix":  [],
                }

            g = groups[key]
            g["err"].extend(err_lines)
            g["page"].extend(page_lines)
            g["fix"].extend(fix_lines)

    absorb(ws1, last_row_1, 1)
    absorb(ws2, last_row_2, 2)
    absorb(ws4, last_row_4, 4)

    return list(groups.values())

# =========================
# 공개 API (GUI/모듈에서 호출)
# =========================
def process_error_list(src_path: str, on_progress=None) -> Dict:
    """src_path를 읽어 자동분류 파일을 생성하고 결과 정보를 반환합니다.

    반환 예시:
      {"src_path": str, "dst_path": str, "total_rows": int, "copied_without_processing": bool}
    """
    def progress(pct: int, msg: str):
        if on_progress is not None:
            try:
                on_progress(pct, msg)
            except Exception:
                pass

    if not src_path:
        raise ValueError("파일 경로가 비어 있습니다.")
    if not os.path.exists(src_path):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {src_path}")
    if not src_path.lower().endswith('.xlsx'):
        raise ValueError(f"Excel 파일(.xlsx)만 지원됩니다: {src_path}")

    base, ext = os.path.splitext(src_path)
    dst_path = f"{base}_자동분류{ext}"

    progress(5, "파일 로딩 중")
    try:
        wb = openpyxl.load_workbook(src_path)
    except Exception as e:
        raise RuntimeError(f"Excel 파일을 열 수 없습니다: {e}")

    missing_sheets = [s for s in SHEETS if s not in wb.sheetnames]
    if missing_sheets:
        raise ValueError(
            f"다음 시트를 찾을 수 없습니다: {', '.join(missing_sheets)} | 현재 시트: {', '.join(wb.sheetnames)}"
        )

    ws1, ws2, ws3, ws4 = [wb[s] for s in SHEETS]

    progress(10, "1차·2차·품질점검 시트 데이터 병합 중")
    merged = merge_inputs_from_ws1_ws2_ws4(ws1, ws2, ws4)
    total_rows = len(merged)

    if total_rows == 0:
        # 입력이 비어 있으면 원본 그대로 저장 후 종료
        try:
            wb.save(dst_path)
        except Exception as e:
            raise RuntimeError(f"파일 저장에 실패했습니다: {e}")
        return {
            "src_path": src_path,
            "dst_path": dst_path,
            "total_rows": 0,
            "copied_without_processing": True,
        }

    new_last_row = DATA_START_ROW + total_rows - 1

    # 4개 시트의 데이터 영역(8행~원래 max_row 또는 new_last_row 중 큰 값)을 모두 클리어
    progress(20, "기존 데이터 영역 클리어 중")
    for ws in (ws1, ws2, ws3, ws4):
        clear_to = max(ws.max_row, new_last_row)
        for r in range(DATA_START_ROW, clear_to + 1):
            for c in range(1, 16):  # A~O
                ws.cell(r, c).value = None

    progress(30, f"오류 항목 자동분류 및 분배 중 (총 {total_rows}행)")

    # 시트별 누적 채움 카운터. 2·3·4차는 "행 위치 기준 페이스(SHEET_TARGET_RATIO × 진행
    # 행수)"에 맞춰 분배하여 처음부터 마지막 행까지 고르게 분포시킨다. 1차는 상한 없는
    # catch-all. (전체 상한을 쓰면 앞쪽 행이 쿼터를 모두 소진해 뒤쪽 행이 비는 현상 발생)
    sheet_filled = {1: 0, 2: 0, 3: 0, 4: 0}

    for idx, mrow in enumerate(merged):
        r = DATA_START_ROW + idx

        # 모든 시트에 A~K 동일하게 기록 (병합 결과의 캡처값)
        for ws in (ws1, ws2, ws3, ws4):
            for c, val in enumerate(mrow["a_to_k"], start=1):
                ws.cell(r, c).value = val

        types = [normalize_error_type(x) for x in mrow["err"]]
        pages = mrow["page"]
        fixes = mrow["fix"]
        L = min(len(types), len(pages), len(fixes))

        if L == 0:
            # 해당 제어번호에 분배할 오류 항목이 없음 - A~K만 기록하고 다음으로
            continue

        raw_items: List[Dict] = []
        for i in range(L):
            # '/' 분리는 분배 이후(렌더링) 단계로 미룬다. 여기서는 fix 한 줄 = 항목 1개로 다룬다.
            # 원본 문자열(raw_fix)은 보관했다가 분배 후 렌더링에서 '/' 분리·diff에 사용한다.
            cleaned = fixes[i].replace("⌄", " ").replace("⌴", " ").replace("­", "")  # 소프트하이픈(U+00AD) 제거
            # 중복판정용 before/after는 '/' 병합된 깨끗한 값으로 (렌더링과 동일 기준)
            before, after, found_arrow = merged_before_after(cleaned, types[i], pages[i])

            err = types[i]
            # 오탈자 중 번호체계 괄호↔온점 변환('[...]'↔'....', 양방향)이면 '책갈피수정'으로 재분류.
            # ('/'로 묶인 경우 모든 조각이 괄호↔온점 변환일 때만)
            if err == "오탈자":
                sub_parsed = [parse_fix(s, err, pages[i]) for s in split_multi_fix(cleaned)]
                if sub_parsed and all(is_bracket_dot_change(b, a) for b, a, _ in sub_parsed):
                    err = "책갈피수정"

            raw_items.append({
                "arrow": found_arrow,
                "err": err,
                "page": pages[i],
                "before": before,
                "after": after,
                "raw_fix": cleaned,
            })

        deduplicated_items = remove_duplicates(raw_items)

        kept_per_kind: Dict[str, int] = {}
        per_kind_capped: List[Dict] = []
        for it in deduplicated_items:
            k = it["err"]
            cnt = kept_per_kind.get(k, 0)
            if cnt < MAX_PER_ERR_KIND:
                per_kind_capped.append(it)
                kept_per_kind[k] = cnt + 1

        capped = per_kind_capped[:MAX_PER_ROW_TOTAL]

        items_1: List[Dict] = []
        items_2: List[Dict] = []
        items_3: List[Dict] = []
        items_4: List[Dict] = []

        # 쿼터 기반 분배(한 항목 = 한 시트, 행 수 기준 비율 수렴).
        pool = list(capped)
        late_bins = {2: items_2, 3: items_3, 4: items_4}
        filled_this_row: set = set()

        # ① 1차 점검: 항목이 있으면 첫 항목을 우선 배정(주 점검 시트, 상한 없음).
        if pool:
            items_1.append(pool.pop(0))
            sheet_filled[1] += 1
            filled_this_row.add(1)

        # ② 2·3·4차: 행 위치 기준 페이스(SHEET_TARGET_RATIO × 진행 행수)에 따라 분배.
        #    목표 페이스보다 뒤처진 시트부터 한 항목씩 채워, 마지막 행까지 고르게 분포시킨다.
        #    각 시트는 이 행에서 최대 1회. 3·4차는 오탈자/띄어쓰기 항목만 받음.
        pos = idx + 1  # 전체 데이터 행 중 현재 행의 1-based 위치
        while pool:
            best_s = None
            best_deficit = None
            for s in (2, 3, 4):
                if s in filled_this_row:
                    continue
                paced_target = SHEET_TARGET_RATIO[s] * pos  # 이 위치까지의 목표 누적치
                if sheet_filled[s] >= paced_target:
                    continue  # 이미 페이스만큼 채움 → 이후 행을 위해 양보
                if s in (3, 4) and not any(it["err"] in ALLOWED_TYPES_LATE for it in pool):
                    continue
                deficit = paced_target - sheet_filled[s]
                if best_deficit is None or deficit > best_deficit:
                    best_s, best_deficit = s, deficit
            if best_s is None:
                break
            if best_s in (3, 4):
                pick = next(i for i, it in enumerate(pool) if it["err"] in ALLOWED_TYPES_LATE)
            else:
                pick = 0
            late_bins[best_s].append(pool.pop(pick))
            sheet_filled[best_s] += 1
            filled_this_row.add(best_s)

        # ③ 남은 잉여 항목은 모두 1차 점검으로(데이터 손실 없이 흡수).
        if pool:
            if 1 not in filled_this_row:
                sheet_filled[1] += 1
            items_1.extend(pool)
            pool = []

        # L/M/N 값만 기록 (정렬 등 스타일은 마지막 일괄 적용 단계에서 처리)
        # 분배 이후 각 항목을 '/' 기준으로 펼쳐 L/M/N 줄 수를 일치시킨다.
        for ws, items in ((ws1, items_1), (ws2, items_2), (ws3, items_3), (ws4, items_4)):
            if not items:
                continue  # 이미 위에서 None으로 클리어됨
            units = expand_fix_units(items)
            if not units:
                continue
            # L(오류항목)·M(페이지)는 '항목'마다 한 번만 기록.
            # '/'로 분리된 같은 항목의 이어지는 줄은 빈칸 → 같은 항목임이 보이고,
            # 서로 다른 항목은 각자 값이 찍혀 수정내역끼리 구분이 된다.
            ws.cell(r, COL_ERR).value  = "\n".join(u["err"] if u["item_start"] else "" for u in units)
            ws.cell(r, COL_PAGE).value = "\n".join(str(u["page"]) if u["item_start"] else "" for u in units)
            apply_rich_diff_for_bidir(ws.cell(r, COL_FIX), units)

    # === 마지막 일괄 스타일 적용 단계 ===
    progress(80, "스타일 적용 중 (폰트·정렬·행높이·테두리)")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    base_font   = Font(name=FONT_NAME, size=FONT_SIZE)               # 맑은 고딕 10pt
    center_mid  = Alignment(wrap_text=True, horizontal='center', vertical='center')  # A~C, O
    center_top  = Alignment(wrap_text=True, horizontal='center', vertical='top')     # L, M
    top_align   = Alignment(wrap_text=True, vertical='top')          # N

    # 모든 시트 A~O열, 8행~마지막행에 폰트/테두리/정렬 일괄 적용
    for sheet in (ws1, ws2, ws3, ws4):
        for row_num in range(DATA_START_ROW, new_last_row + 1):
            # 행 높이 자동 맞춤
            sheet.row_dimensions[row_num].auto_size = True
            for col_num in range(1, 16):  # A~O
                cell = sheet.cell(row_num, col_num)
                cell.font = base_font          # 맑은 고딕 10pt
                cell.border = thin_border
                if col_num in (1, 2, 3, 15):   # A~C, O → 가운데맞춤(가로·세로)
                    cell.alignment = center_mid
                elif col_num in (COL_ERR, COL_PAGE):  # L, M → 가로 가운데 + 위쪽
                    cell.alignment = center_top
                elif col_num == COL_FIX:        # N → 위쪽맞춤
                    cell.alignment = top_align

    progress(95, "파일 저장 중")
    try:
        wb.save(dst_path)
    except Exception as e:
        raise RuntimeError(f"파일 저장에 실패했습니다: {e}")

    progress(100, "완료")
    return {
        "src_path": src_path,
        "dst_path": dst_path,
        "total_rows": total_rows,
        "copied_without_processing": False,
    }

# =========================
# CLI 전용 실행부
# =========================
def main():
    src_path, dst_path = get_file_paths()
    print(f"📂 파일 로딩 중: {os.path.basename(src_path)}")

    try:
        result = process_error_list(src_path)
        total_rows = result.get("total_rows", 0)
        copied = result.get("copied_without_processing", False)

        if copied:
            print(f"ℹ️  처리할 데이터가 없습니다.")
            print(f"✅ 원본 파일을 복사하여 저장했습니다: {os.path.basename(result['dst_path'])}")
            input("아무 키나 눌러 종료하세요...")
            sys.exit(0)

        print("=" * 60)
        print("🎉 자동분류 완료!")
        print("=" * 60)
        print(f"📁 입력 파일: {os.path.basename(result['src_path'])}")
        print(f"📁 출력 파일: {os.path.basename(result['dst_path'])}")
        print(f"📊 처리된 행수: {total_rows}행")
        print()
        print("🔸 분류 결과:")
        print("  - 중복 제거: 85% 이상 유사한 항목 자동 제거")
        print(f"  - 1차 점검: 주 점검 + 잉여 항목 흡수 (상한 없음)")
        print(f"  - 2차 점검: 목표 약 {int(SHEET_TARGET_RATIO[2]*100)}% 행 (쿼터)")
        print(f"  - 3차 점검: 목표 약 {int(SHEET_TARGET_RATIO[3]*100)}% 행 (오탈자/띄어쓰기만)")
        print(f"  - 품질 점검: 목표 약 {int(SHEET_TARGET_RATIO[4]*100)}% 행 (오탈자/띄어쓰기만)")
        print("  - 행 높이: 8행부터 자동 맞춤 적용")
        print("  - 테두리: 8행부터 O열까지 모든 셀에 적용")
        print("=" * 60)
    except Exception as e:
        print(f"❌ 오류: {e}")
        input("아무 키나 눌러 종료하세요...")
        sys.exit(1)

    input("아무 키나 눌러 종료하세요...")


if __name__ == "__main__":
    main()