# -*- coding: utf-8 -*-
"""
scrape_distr_pipeline/diff_core.py  ── (선택 단계) 수정내역 빨간색 diff 강조

VBA 매크로 `Module1.HighlightTyposInSelection`의 동작을 Python으로 포팅한 모듈.
L/N열이 이미 기록된 엑셀 파일을 받아, 4개 시트의 N열 데이터에 대해
오타/차이 부분을 빨간색으로 표시(구두점은 Bold) Rich Text를 적용합니다.

처리 절차 (셀 단위):
    1) 특수공백 ChrW(9012)/ChrW(9251) → 일반 공백 치환
    2) 줄 단위 세그먼트 빌드 (단일 / 페어 / n.X→Y 인라인 / 다음 줄 화살표)
       - 좌/우 동일 시 "띄어쓰기 또는 단계 확인 필요" 부가
       - 선두 "→ 추가 ..." 패턴 정규화 ("추가 ...")
    3) 세그먼트별 페인트
       - 페어: 좌/우 비교 → 차이 글자 빨강 + 인접 경계 1글자 빨강
       - 단일: 문장 시작/끝의 "추가/삭제" 라벨만 빨강
       - 구두점(. , : ; ·)은 Bold 추가

공개 함수
    run_diff_highlight(excel_file, log_callback)
        → 4개 시트 N열에 Rich Text diff 강조 적용 후 저장
    patch_xml_space_preserve(excel_file, log_callback)
        → openpyxl 3.1.5 버그 우회 (xml:space=preserve 보강)
"""

import os
import re
import shutil
import zipfile
from typing import Callable, List, Optional

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from common.constants import DATA_START_ROW, SHEETS
from common.utils import (
    default_log_callback,
    normalize_error_type,
    split_fix_lines,
    split_lines,
)

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------
_COL_ERR = 12   # L: 오류항목
_COL_FIX = 14   # N: 수정내역

# 오류항목 종류별 처리 모드
_FULL_DIFF_TYPES = {"오탈자", "띄어쓰기", "책갈피수정"}   # 화살표 줄바꿈 + 빨간색
_SKIP_TYPES      = {"링크", "단계"}                       # 처리 제외

_INLINE_FONT_NAME    = "맑은 고딕"
_INLINE_FONT_SIZE    = 10
_INLINE_FONT_CHARSET = 129   # Korean (Hangul)
_INLINE_FONT_FAMILY  = 2     # Swiss / sans-serif

_PUNCT_CHARS    = {".", ",", ":", ";", "·"}
_WS_CHARS       = {" ", "\t"}
_SPECIAL_SPACES = (chr(9012), chr(9251))   # Counterbore, Open Box 기호

_RE_NDOT_NUM  = re.compile(r"n\.(\d+)\s*(→|->|=>)\s*(\d+)")
_RE_ADDITION  = re.compile(r"(^\s*(추가|삭제)\s+\S)|(\S\s+(추가|삭제)\s*$)")


def _inline_font(color: str, bold: bool = False) -> InlineFont:
    return InlineFont(
        rFont   = _INLINE_FONT_NAME,
        charset = _INLINE_FONT_CHARSET,
        family  = _INLINE_FONT_FAMILY,
        sz      = _INLINE_FONT_SIZE,
        color   = color,
        b       = bold,
    )


# ---------------------------------------------------------------------------
# 판정 헬퍼
# ---------------------------------------------------------------------------

def _is_ws(ch: str) -> bool:
    return ch in _WS_CHARS


def _is_punct(ch: str) -> bool:
    return ch in _PUNCT_CHARS


def _starts_with_ndot(s: str) -> bool:
    return len(s) >= 2 and s[:2] == "n."


def _skip_ws_forward(s: str, i: int) -> int:
    """0-based: i부터 연속 공백 건너뛴 첫 비공백 인덱스를 반환."""
    n = len(s)
    while i < n and _is_ws(s[i]):
        i += 1
    return i


# ---------------------------------------------------------------------------
# 텍스트 정규화
# ---------------------------------------------------------------------------

def _normalize_leading_arrow_addition(s: str) -> str:
    """선두 "  →   추가 안내" → "추가 안내" 치환."""
    n = len(s)
    i = 0
    while i < n and s[i] in (" ", "\t"):
        i += 1
    if i >= n:
        return s
    if s[i] == "→":
        arrow_len = 1
    elif i + 1 < n and s[i:i+2] == "->":
        arrow_len = 2
    else:
        return s
    i += arrow_len
    while i < n and s[i] in (" ", "\t"):
        i += 1
    if s[i:i+2] == "추가":
        return "추가" + s[i+2:]
    return s


# ---------------------------------------------------------------------------
# 세그먼트 빌드
# ---------------------------------------------------------------------------

def _add_seg(segs: list, kind: str, disp: str,
             cmp_l: str = "", cmp_r: str = "", has_delim: bool = False,
             inline_l_start: int = 0, inline_r_start: int = 0) -> None:
    segs.append({
        "kind":         kind,
        "disp":         disp,
        "cmpL":         cmp_l,
        "cmpR":         cmp_r,
        "hasDelim":     has_delim,
        "inlineLStart": inline_l_start,
        "inlineRStart": inline_r_start,
    })


def _has_differences(left_cmp: str, right_cmp: str) -> bool:
    iL = iR = 0
    nL, nR = len(left_cmp), len(right_cmp)
    while iL < nL and iR < nR:
        if left_cmp[iL] == right_cmp[iR]:
            iL += 1; iR += 1
        else:
            return True
    return iL < nL or iR < nR


def _find_arrow(line_text: str) -> tuple:
    """(delim_pos, delim_len) 반환. 없으면 (-1, 0)."""
    p = line_text.find("→")
    if p >= 0:
        return p, 1
    p = line_text.find("->")
    if p >= 0:
        return p, 2
    p = line_text.find("=>")
    if p >= 0:
        return p, 2
    return -1, 0


def _build_segments_from_lines(lines: List[str], i: int, segs: list) -> int:
    """lines[i]를 처리하고, BuildSegmentsFromLines 종료 시점의 i를 반환.

    호출자가 추가로 +1 하므로, 다음 줄까지 소비한 경우 i+1을 반환.
    """
    line_text = _normalize_leading_arrow_addition(lines[i])

    # 1) "n.숫자 → 숫자" 인라인 패턴
    m = _RE_NDOT_NUM.search(line_text)
    if m:
        num_l, arrow_txt, num_r = m.group(1), m.group(2), m.group(3)
        match_text = m.group(0)
        base_pos   = m.start() + 1
        rel_l      = match_text.find(num_l) + 1
        rel_arrow  = match_text.find(arrow_txt) + 1
        rel_r      = match_text.find(num_r, rel_arrow + len(arrow_txt) - 1) + 1
        abs_l      = base_pos + rel_l - 1
        abs_r      = base_pos + rel_r - 1
        _add_seg(segs, "inline_num_pair", line_text,
                 "n." + num_l, num_r, True,
                 inline_l_start=abs_l - 2, inline_r_start=abs_r)
        return i

    # 2) 화살표 위치
    delim_pos, delim_len = _find_arrow(line_text)

    # CASE A: 화살표가 다음 줄 시작에 있는 경우
    if delim_pos < 0:
        if i < len(lines) - 1:
            next_line = lines[i + 1].lstrip(" \t")
            next_delim_len = 0
            if next_line.startswith("→"):
                next_delim_len = 1
            elif next_line.startswith("->") or next_line.startswith("=>"):
                next_delim_len = 2

            if next_delim_len > 0:
                left_raw   = line_text
                delim_text = next_line[:next_delim_len]
                right_raw  = next_line[next_delim_len:]
                l_trimmed  = left_raw.strip()
                r_trimmed  = right_raw.strip()
                off_l = left_raw.find(l_trimmed) + 1
                off_r = next_delim_len + (right_raw.find(r_trimmed) + 1)

                _add_seg(segs, "pair_left",  left_raw,
                         l_trimmed, r_trimmed, True,
                         inline_l_start=off_l, inline_r_start=0)
                _add_seg(segs, "pair_right", delim_text + right_raw,
                         l_trimmed, r_trimmed, True,
                         inline_l_start=0, inline_r_start=off_r)
                return i + 1
        _add_seg(segs, "single", line_text)
        return i

    # CASE B: 화살표가 한 줄 중간에 있는 경우
    left_raw   = line_text[:delim_pos]
    right_raw  = line_text[delim_pos + delim_len:]
    delim_text = line_text[delim_pos:delim_pos + delim_len]

    if _starts_with_ndot(left_raw):
        _add_seg(segs, "single", line_text, has_delim=True)
        return i

    # 구분자 오른쪽이 비어 있으면 다음 줄을 내용으로 가져옴
    if right_raw.strip() == "" and i < len(lines) - 1:
        i += 1
        right_raw = lines[i]

    l_trimmed = left_raw.strip()
    r_trimmed = right_raw.strip()
    off_l = left_raw.find(l_trimmed) + 1
    # pair_right disp는 선행 공백을 제거 — 화살표 직후 공백이
    # 줄바꿈(full_diff) 시 다음 줄 앞에 끌려가는 문제 방지.
    right_disp = right_raw.lstrip()
    off_r = 1

    if _has_differences(l_trimmed, r_trimmed):
        _add_seg(segs, "pair_left",  left_raw + delim_text,
                 l_trimmed, r_trimmed, True,
                 inline_l_start=off_l, inline_r_start=0)
        _add_seg(segs, "pair_right", right_disp,
                 l_trimmed, r_trimmed, True,
                 inline_l_start=0, inline_r_start=off_r)
    else:
        _add_seg(segs, "single",
                 left_raw + delim_text + " 띄어쓰기 또는 단계 확인 필요",
                 has_delim=True)
    return i


# ---------------------------------------------------------------------------
# 페인트
# ---------------------------------------------------------------------------

def _paint_char(s: str, is_red: List[bool], is_bold: List[bool],
                abs_pos_1based: int) -> None:
    idx = abs_pos_1based - 1
    if 0 <= idx < len(s):
        is_red[idx] = True
        if _is_punct(s[idx]):
            is_bold[idx] = True


def _paint_span(s: str, is_red: List[bool], is_bold: List[bool],
                abs_start_1based: int, span_len: int) -> None:
    total = len(s)
    if abs_start_1based < 1 or span_len <= 0 or abs_start_1based > total:
        return
    if abs_start_1based + span_len - 1 > total:
        span_len = total - abs_start_1based + 1
    for k in range(abs_start_1based, abs_start_1based + span_len):
        idx = k - 1
        is_red[idx] = True
        if _is_punct(s[idx]):
            is_bold[idx] = True


def _paint_pair(s: str, is_red: List[bool], is_bold: List[bool],
                left_cmp: str, right_cmp: str,
                left_start_1based: int, right_start_1based: int) -> None:
    """좌/우 페어 비교 → 차이 부분 + 인접 경계 글자 빨간색."""
    nL, nR = len(left_cmp), len(right_cmp)
    iL = iR = 0   # 0-based

    while iL < nL and iR < nR:
        chL, chR = left_cmp[iL], right_cmp[iR]
        if chL == chR:
            iL += 1; iR += 1
        elif _is_ws(chL) and not _is_ws(chR):
            nxt_l = _skip_ws_forward(left_cmp, iL)
            if iL >= 1:
                _paint_char(s, is_red, is_bold, left_start_1based + (iL - 1))
            if nxt_l < nL:
                _paint_char(s, is_red, is_bold, left_start_1based + nxt_l)
            if iR >= 1:
                _paint_char(s, is_red, is_bold, right_start_1based + (iR - 1))
            _paint_char(s, is_red, is_bold, right_start_1based + iR)
            iL = nxt_l
        elif not _is_ws(chL) and _is_ws(chR):
            nxt_r = _skip_ws_forward(right_cmp, iR)
            if iL >= 1:
                _paint_char(s, is_red, is_bold, left_start_1based + (iL - 1))
            _paint_char(s, is_red, is_bold, left_start_1based + iL)
            if iR >= 1:
                _paint_char(s, is_red, is_bold, right_start_1based + (iR - 1))
            if nxt_r < nR:
                _paint_char(s, is_red, is_bold, right_start_1based + nxt_r)
            iR = nxt_r
        else:
            _paint_char(s, is_red, is_bold, left_start_1based + iL)
            _paint_char(s, is_red, is_bold, right_start_1based + iR)
            iL += 1; iR += 1

    # 잔여(길이 차) 처리
    if iL < nL:
        _paint_span(s, is_red, is_bold, left_start_1based + iL, nL - iL)
    if iR < nR:
        _paint_span(s, is_red, is_bold, right_start_1based + iR, nR - iR)


def _highlight_addition(s: str, is_red: List[bool], is_bold: List[bool],
                        line_text: str, abs_start_1based: int) -> None:
    """문장 시작/끝의 "추가" 또는 "삭제" 라벨만 빨간색 처리."""
    for m in _RE_ADDITION.finditer(line_text):
        matched = m.group(0)
        if "추가" in matched:
            word = "추가"
        elif "삭제" in matched:
            word = "삭제"
        else:
            continue
        word_pos = m.start() + matched.find(word) + 1   # 1-based in line_text
        _paint_span(s, is_red, is_bold,
                    abs_start_1based + word_pos - 1, len(word))


def _paint_segments(s: str, segs: list, pair_sep_len: int,
                    is_red: List[bool], is_bold: List[bool]) -> None:
    """세그먼트별로 페인트. pair_left/pair_right 사이 구분자 길이는 가변."""
    cursor = 1   # 1-based 위치
    n = len(segs)
    i = 0
    while i < n:
        seg = segs[i]
        if seg["kind"] == "pair_left":
            left_disp  = seg["disp"]
            right_seg  = segs[i + 1]
            right_disp = right_seg["disp"]
            left_start  = cursor + (seg["inlineLStart"] - 1)
            right_start = (cursor + len(left_disp) + pair_sep_len) + (right_seg["inlineRStart"] - 1)
            _paint_pair(s, is_red, is_bold,
                        seg["cmpL"], seg["cmpR"],
                        left_start, right_start)
            cursor += len(left_disp) + pair_sep_len + len(right_disp) + (1 if i + 2 < n else 0)
            i += 2
        else:
            _highlight_addition(s, is_red, is_bold, seg["disp"], cursor)
            cursor += len(seg["disp"]) + (1 if i < n - 1 else 0)
            i += 1


# ---------------------------------------------------------------------------
# Rich Text 빌드
# ---------------------------------------------------------------------------

def _build_rich_text(text: str,
                     is_red: List[bool], is_bold: List[bool]) -> CellRichText:
    crt = CellRichText()
    if not text:
        return crt
    n = len(text)
    start    = 0
    cur_red  = is_red[0]
    cur_bold = is_bold[0]
    for k in range(1, n):
        if is_red[k] != cur_red or is_bold[k] != cur_bold:
            color = "FFFF0000" if cur_red else "FF000000"
            crt.append(TextBlock(text=text[start:k],
                                 font=_inline_font(color, cur_bold)))
            start    = k
            cur_red  = is_red[k]
            cur_bold = is_bold[k]
    color = "FFFF0000" if cur_red else "FF000000"
    crt.append(TextBlock(text=text[start:n],
                         font=_inline_font(color, cur_bold)))
    return crt


# ---------------------------------------------------------------------------
# 셀 처리
# ---------------------------------------------------------------------------

def _cell_text(value) -> str:
    """openpyxl 셀 값(plain str / CellRichText / None)에서 평문 추출."""
    if value is None:
        return ""
    if isinstance(value, CellRichText):
        parts = []
        for t in value:
            if isinstance(t, str):
                parts.append(t)
            else:
                parts.append(getattr(t, "text", str(t)))
        return "".join(parts)
    return str(value)


def _resolve_mode(orig_err: str) -> str:
    """오류 종류 → 처리 모드("skip" | "full_diff" | "red_only")."""
    norm = normalize_error_type(orig_err)
    if norm in _SKIP_TYPES:
        return "skip"
    if norm in _FULL_DIFF_TYPES:
        return "full_diff"
    return "red_only"


def _process_item(item_text: str, mode: str) -> tuple:
    """단일 수정내역 항목을 모드에 따라 처리. (text, is_red, is_bold) 반환."""
    if mode == "skip" or not item_text:
        n = len(item_text)
        return item_text, [False] * n, [False] * n

    # 1) 특수공백 → 일반 공백
    text = item_text
    for sp in _SPECIAL_SPACES:
        text = text.replace(sp, " ")
    text = text.replace("\r\n", "\n")

    # 2) 세그먼트 빌드
    lines = text.split("\n")
    segs: list = []
    i = 0
    while i < len(lines):
        i = _build_segments_from_lines(lines, i, segs)
        i += 1

    if not segs:
        n = len(text)
        return text, [False] * n, [False] * n

    # 3) 모드별 disp 결합 - pair_left/right 사이 구분자만 가변
    #   full_diff: "\n" (줄바꿈)
    #   red_only:  " "  (한 줄 유지, pair_right disp는 선행 공백 제거됨)
    pair_sep = "\n" if mode == "full_diff" else " "

    pieces = []
    for j, seg in enumerate(segs):
        if j > 0:
            prev = segs[j - 1]
            if prev["kind"] == "pair_left" and seg["kind"] == "pair_right":
                pieces.append(pair_sep)
            else:
                pieces.append("\n")
        pieces.append(seg["disp"])
    new_text = "".join(pieces)

    is_red  = [False] * len(new_text)
    is_bold = [False] * len(new_text)

    # 4) 페인트
    _paint_segments(new_text, segs, len(pair_sep), is_red, is_bold)

    return new_text, is_red, is_bold


def _process_cell_value(l_value, n_value) -> Optional[CellRichText]:
    """L/N 셀 값을 받아 항목별 종류에 따라 다르게 diff 적용한 CellRichText 반환.

    - 링크/단계: 그대로 통과 (페인트 X)
    - 오탈자/띄어쓰기/책갈피수정: 화살표 줄바꿈 + 빨간색
    - 그 외 (추가/삭제 계열 등): 줄바꿈 없이 빨간색만
    """
    items = split_fix_lines(n_value)
    if not items:
        return None
    types = split_lines(l_value)

    modes = [
        _resolve_mode(types[k] if k < len(types) else "")
        for k in range(len(items))
    ]

    # 한 행에 항목이 2개 이상이면 줄바꿈 스킵 (full_diff → red_only)
    # 단, 링크/단계는 그대로 skip 유지.
    if len(items) >= 2:
        modes = ["red_only" if m == "full_diff" else m for m in modes]

    # 전부 skip이면 셀 변경 불필요
    if all(m == "skip" for m in modes):
        return None

    # 항목별 처리
    item_results = [_process_item(items[k], modes[k]) for k in range(len(items))]

    # 항목 간 \n 연결
    text_parts: List[str] = []
    all_red:  List[bool] = []
    all_bold: List[bool] = []
    for idx, (text, red, bold) in enumerate(item_results):
        if idx > 0:
            text_parts.append("\n")
            all_red.append(False)
            all_bold.append(False)
        text_parts.append(text)
        all_red.extend(red)
        all_bold.extend(bold)

    new_text = "".join(text_parts)
    if not new_text:
        return None
    return _build_rich_text(new_text, all_red, all_bold)


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def run_diff_highlight(excel_file: str, log_callback: Callable = None) -> str:
    """엑셀 파일 4개 시트의 N열에 매크로 기반 diff 강조를 적용합니다.

    Args:
        excel_file:   처리할 엑셀 파일 경로 (.xlsx 또는 .xlsm)
        log_callback: log_callback(status: str, message: str) 패턴

    Returns:
        처리된 파일 경로 (입력 경로와 동일)

    Raises:
        FileNotFoundError: 파일이 존재하지 않을 때
        ValueError:        지원하지 않는 확장자 또는 시트 누락 시
        RuntimeError:      파일 열기·저장 실패 시
    """
    log_callback = log_callback or default_log_callback

    if not excel_file:
        raise ValueError("파일 경로가 비어 있습니다.")
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {excel_file}")
    _, ext = os.path.splitext(excel_file)
    if ext.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(f".xlsx 또는 .xlsm 파일만 지원됩니다: {excel_file}")

    log_callback("info", f"📍 '{excel_file}' 열기 중...")
    is_macro = ext.lower() == ".xlsm"
    try:
        wb = load_workbook(excel_file, keep_vba=is_macro, rich_text=True)
    except Exception as e:
        raise RuntimeError(f"Excel 파일을 열 수 없습니다: {e}") from e

    missing = [s for s in SHEETS if s not in wb.sheetnames]
    if missing:
        wb.close()
        raise ValueError(
            f"다음 시트를 찾을 수 없습니다: {', '.join(missing)} "
            f"| 현재 시트: {', '.join(wb.sheetnames)}"
        )

    try:
        applied = 0
        for sheet_name in SHEETS:
            ws = wb[sheet_name]
            for row_idx in range(DATA_START_ROW, ws.max_row + 1):
                l_value = ws.cell(row_idx, _COL_ERR).value
                n_cell  = ws.cell(row_idx, _COL_FIX)
                rich = _process_cell_value(l_value, n_cell.value)
                if rich is None:
                    continue
                n_cell.value = rich
                applied += 1

        log_callback("info", f"📍 {applied}개 셀 Rich Text diff 적용 완료. 저장 중...")
        try:
            wb.save(excel_file)
        except Exception as e:
            raise RuntimeError(f"파일 저장에 실패했습니다: {e}") from e
    finally:
        wb.close()

    log_callback("info", f"✅ Rich Text diff 강조 완료. → '{excel_file}'")
    return excel_file


# ---------------------------------------------------------------------------
# XML 후처리 (openpyxl 3.1.5 버그 우회)
# ---------------------------------------------------------------------------

_T_ELEMENT_RE = re.compile(r"<t([^>]*)>([^<]*)</t>")


def _needs_preserve(content: str) -> bool:
    if not content:
        return False
    if content[0] in " \n\r\t" or content[-1] in " \n\r\t":
        return True
    return "\n" in content or "\r" in content


def _patch_t_elements(xml: str) -> str:
    def replace(m: re.Match) -> str:
        attrs, content = m.group(1), m.group(2)
        if "xml:space" in attrs or not _needs_preserve(content):
            return m.group(0)
        return f'<t{attrs} xml:space="preserve">{content}</t>'
    return _T_ELEMENT_RE.sub(replace, xml)


def patch_xml_space_preserve(excel_file: str, log_callback: Callable = None) -> str:
    """엑셀 파일의 worksheet XML에서 누락된 xml:space='preserve'를 보강합니다.

    openpyxl 3.1.5는 단일 공백 TextBlock 등 일부 인라인 문자열에 대해
    xml:space='preserve'를 빠뜨리는 버그가 있어, Excel이 파일을 열 때
    "복구된 레코드 — 문자열 속성" 경고를 띄움. 저장 후 zip 내부의 worksheet
    xml들을 직접 패치해 해결.
    """
    log_callback = log_callback or default_log_callback

    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {excel_file}")

    log_callback("info", "📍 worksheet XML 후처리 (xml:space=preserve 보강) 중...")

    tmp_path = excel_file + ".tmp"
    patched_count = 0
    try:
        with zipfile.ZipFile(excel_file, "r") as src, \
             zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as dst:
            for item in src.infolist():
                data = src.read(item.filename)
                if (item.filename.startswith("xl/worksheets/")
                        and item.filename.endswith(".xml")):
                    xml = data.decode("utf-8")
                    new_xml = _patch_t_elements(xml)
                    if new_xml != xml:
                        patched_count += 1
                    data = new_xml.encode("utf-8")
                dst.writestr(item, data)
        shutil.move(tmp_path, excel_file)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise

    log_callback("info", f"✅ {patched_count}개 worksheet 패치 완료.")
    return excel_file
