# -*- coding: utf-8 -*-
"""
scrape_distr_pipeline/dist_core.py  ── 2단계: 오류항목 자동분배

scrape_core.run_scrape()가 수집한 오류항목을 분배 규칙에 따라
4개 시트(1차~품질점검)에 배분하고 엑셀 파일에 기록합니다.

공개 함수
    run_distribute(rows_list, parsed_items_list, excel_file, log_callback)
        → 원본 엑셀 복사 → A~H 공통정보 기록 → L/M/N 오류항목 분배
        → _자동분류{원본확장자} 파일로 저장 → 경로 반환

    process_distribution(src_path, log_callback)
        → (독립 실행용) 이미 1차 점검 시트에 L/M/N이 기록된 엑셀 파일을
          받아 분배만 수행합니다. pipeline과 무관하게 단독으로 사용 가능합니다.

분배 규칙
    중복제거 (85% 유사도 기준) → 종류별 최대 5개 → 행당 최대 5개
    1차 점검: 1개 / 2차 점검: 나머지 1개
    3차 점검: 오탈자·띄어쓰기 우선 1개, 없으면 2차에 추가
    품질 점검: 오탈자·띄어쓰기 우선 1개, 없으면 2차에 추가
"""

import os
import shutil
from difflib import SequenceMatcher
from typing import Callable, Dict, List, Tuple

from openpyxl import load_workbook

from common.constants import (
    ARROW,
    COMMON_COLUMN_INDEX,
    DATA_START_ROW,
    MIN_SHEET_COUNT,
    OUTPUT_SUFFIX,
    SHEETS,
)
from common.utils import default_log_callback, get_val, normalize_error_type, parse_fix_text, split_lines, split_fix_lines

# ---------------------------------------------------------------------------
# 상수 (dist_core 전용)
# ---------------------------------------------------------------------------
_COL_ERR  = 12   # L: 오류항목
_COL_PAGE = 13   # M: 페이지번호
_COL_FIX  = 14   # N: 수정내역

_ALLOWED_TYPES_LATE   = {"오탈자", "띄어쓰기"}   # 3차/품질 시트에서 우선 배치
_MAX_PER_ERR_KIND     = 5
_MAX_PER_ROW_TOTAL    = 5
_SIMILARITY_THRESHOLD = 0.85


# ---------------------------------------------------------------------------
# 내부 함수
# ---------------------------------------------------------------------------

def _calculate_similarity(item1: Dict, item2: Dict) -> float:
    """두 오류 항목 dict의 유사도를 [0.0, 1.0]으로 반환합니다.

    오류 유형이 다르면 0.0, 같으면 before/after SequenceMatcher 평균.
    """
    if item1["err"] != item2["err"]:
        return 0.0
    before_sim = SequenceMatcher(None, item1["before"], item2["before"]).ratio()
    after_sim  = SequenceMatcher(None, item1["after"],  item2["after"]).ratio()
    return (before_sim + after_sim) / 2.0


def _remove_duplicates(items: List[Dict]) -> List[Dict]:
    """SIMILARITY_THRESHOLD 이상 유사한 항목을 제거한 리스트를 반환합니다."""
    unique: List[Dict] = []
    for current in items:
        if not any(
            _calculate_similarity(current, existing) >= _SIMILARITY_THRESHOLD
            for existing in unique
        ):
            unique.append(current)
    return unique


def _apply_prefix(orig_err: str, after: str) -> str:
    """추가/삭제 계열 항목의 after 텍스트에 '추가 '/'삭제 ' prefix를 부여합니다."""
    if orig_err in ("추가", "책갈피추가") and not after.startswith("추가 "):
        return f"추가 {after}"
    if orig_err in ("삭제", "책갈피삭제") and not after.startswith("삭제 "):
        return f"삭제 {after}"
    return after


def _format_fix_line(item: Dict) -> str:
    """단일 항목을 'before {arrow} after' 또는 'after' plain text로 반환합니다."""
    before = item.get("before") or ""
    after  = _apply_prefix(item.get("orig_err", ""), (item.get("after") or "").lstrip())
    arrow  = item.get("arrow") or ARROW
    return f"{before} {arrow} {after}" if before else after


def _cap_per_kind(items: List[Dict], limit: int) -> List[Dict]:
    """err 종류별로 limit개까지만 남긴 리스트를 반환합니다 (입력 순서 유지)."""
    kept_per_kind: Dict[str, int] = {}
    capped: List[Dict] = []
    for it in items:
        k   = it["err"]
        cnt = kept_per_kind.get(k, 0)
        if cnt < limit:
            capped.append(it)
            kept_per_kind[k] = cnt + 1
    return capped


def _cap_items(raw_items: List[Dict]) -> List[Dict]:
    """중복제거 → 종류별 _MAX_PER_ERR_KIND → 전체 _MAX_PER_ROW_TOTAL 적용."""
    deduped         = _remove_duplicates(raw_items)
    per_kind_capped = _cap_per_kind(deduped, _MAX_PER_ERR_KIND)
    return per_kind_capped[:_MAX_PER_ROW_TOTAL]


def _split_to_sheets(
    capped: List[Dict],
) -> Tuple[List[Dict], List[Dict], List[Dict], List[Dict]]:
    """capped 리스트를 분배 규칙에 따라 4개 시트용 리스트로 나눕니다.

    1차: 최대 1개
    2차: 나머지 중 1개
    3차: _ALLOWED_TYPES_LATE 우선 1개, 없으면 2차에 추가
    품질: _ALLOWED_TYPES_LATE 우선 1개, 없으면 2차에 추가
    """
    items_1: List[Dict] = []
    items_2: List[Dict] = []
    items_3: List[Dict] = []
    items_4: List[Dict] = []

    rest = list(capped)

    if rest:
        idx = next((i for i, it in enumerate(rest) if it["err"] not in _ALLOWED_TYPES_LATE), None)
        items_1.append(rest.pop(idx if idx is not None else 0))

    if rest:
        idx = next((i for i, it in enumerate(rest) if it["err"] not in _ALLOWED_TYPES_LATE), None)
        items_2.append(rest.pop(idx if idx is not None else 0))

    if rest:
        idx = next((i for i, it in enumerate(rest) if it["err"] in _ALLOWED_TYPES_LATE), None)
        if idx is not None:
            items_3.append(rest.pop(idx))
        else:
            items_2.extend(rest[:1])
            rest = rest[1:]

    if rest:
        idx = next((i for i, it in enumerate(rest) if it["err"] in _ALLOWED_TYPES_LATE), None)
        if idx is not None:
            items_4.append(rest.pop(idx))
        else:
            items_2.extend(rest[:1])
            rest = rest[1:]

    return items_1, items_2, items_3, items_4


def _write_lmn(ws, row_idx: int, items: List[Dict]) -> None:
    """ws의 row_idx 행 L/M/N 셀에 items를 기록합니다. 빈 items → None 클리어."""
    err_cell  = ws.cell(row_idx, _COL_ERR)
    page_cell = ws.cell(row_idx, _COL_PAGE)
    fix_cell  = ws.cell(row_idx, _COL_FIX)

    if items:
        err_cell.value  = "\n".join(it["err"]       for it in items)
        page_cell.value = "\n".join(str(it["page"]) for it in items)
        fix_cell.value  = "\n".join(_format_fix_line(it) for it in items)
    else:
        err_cell.value = page_cell.value = fix_cell.value = None


def _write_common_info(sheets: list, rows_list: List[list]) -> None:
    """rows_list의 A~H 공통정보를 4개 시트의 DATA_START_ROW부터 기록합니다."""
    current_rows_position = [DATA_START_ROW] * MIN_SHEET_COUNT
    for row in rows_list:
        for sheet_idx, ws in enumerate(sheets):
            ri = current_rows_position[sheet_idx]
            for gcol, ecol in COMMON_COLUMN_INDEX.items():
                # 빈 값은 None으로 — "" 를 넣으면 openpyxl이
                # <c t="inlineStr" /> (내용 없는 인라인 스트링) 으로 직렬화하여
                # Excel이 "복구된 레코드 — 문자열 속성" 경고를 띄움.
                val = get_val(row, gcol)
                ws.cell(row=ri, column=ecol).value = val if val else None
            current_rows_position[sheet_idx] += 1


def _distribute_and_write(raw_items: List[Dict], excel_row: int, sheets: list) -> None:
    """파싱된 오류항목 리스트를 분배하여 4개 시트에 기록합니다."""
    if not raw_items:
        for ws in sheets[:4]:
            _write_lmn(ws, excel_row, [])
        return

    capped = _cap_items(raw_items)
    items_1, items_2, items_3, items_4 = _split_to_sheets(capped)

    ws1, ws2, ws3, ws4 = sheets[:4]
    _write_lmn(ws1, excel_row, items_1)
    _write_lmn(ws2, excel_row, items_2)
    _write_lmn(ws3, excel_row, items_3)
    _write_lmn(ws4, excel_row, items_4)


def _row_cells_to_raw_items(ws, row_idx: int) -> List[Dict]:
    """ws의 row_idx 행 L/M/N 셀을 읽어 raw_items 리스트로 변환합니다.

    세 컬럼의 줄 수가 다르면 가장 짧은 길이만큼만 처리합니다.
    """
    orig_types = split_lines(ws.cell(row_idx, _COL_ERR).value)
    pages      = split_lines(ws.cell(row_idx, _COL_PAGE).value)
    fixes      = split_fix_lines(ws.cell(row_idx, _COL_FIX).value)
    length     = min(len(orig_types), len(pages), len(fixes))

    raw_items: List[Dict] = []
    for i in range(length):
        found_arrow, before, after = parse_fix_text(fixes[i])
        raw_items.append({
            "arrow":    found_arrow,
            "err":      normalize_error_type(orig_types[i]),
            "orig_err": orig_types[i],
            "page":     pages[i],
            "before":   before,
            "after":    after,
        })
    return raw_items


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def run_distribute(
    rows_list: List[list],
    parsed_items_list: List[List[Dict]],
    excel_file: str,
    log_callback: Callable = None,
) -> str:
    """원본 엑셀을 복사하고 A~H 공통정보와 L/M/N 오류항목을 한 번에 기록합니다.

    scrape_core.run_scrape()의 반환값 (rows_list, parsed_items_list)과
    원본 excel_file을 그대로 전달합니다.

    Args:
        rows_list:         run_scrape()가 반환한 구글 시트 원본 행 데이터 (A~H용)
        parsed_items_list: run_scrape()가 반환한 행별 파싱된 오류항목 (L/M/N용)
        excel_file:        작업 대상 원본 엑셀 파일 (.xlsx 또는 .xlsm)
        log_callback:      log_callback(status: str, message: str) 패턴

    Returns:
        저장된 출력 파일 경로 (_자동분류 접미사)
    """
    log_callback = log_callback or default_log_callback

    if not excel_file:
        raise ValueError("파일 경로가 비어 있습니다.")
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {excel_file}")
    _, ext = os.path.splitext(excel_file)
    if ext.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(f".xlsx 또는 .xlsm 파일만 지원됩니다: {excel_file}")

    if not rows_list:
        log_callback("info", "ℹ️ 분배할 데이터가 없습니다.")
        return excel_file

    # ── 1. 원본 → 출력 파일 복사 ───────────────────────────────────────────
    _, ext   = os.path.splitext(excel_file)
    base     = os.path.splitext(excel_file)[0]
    dst_path = f"{base}{OUTPUT_SUFFIX}{ext}"

    log_callback("info", f"📍 원본을 '{dst_path}'(으)로 복사 중...")
    try:
        shutil.copy2(excel_file, dst_path)
    except Exception as e:
        raise RuntimeError(f"파일 복사에 실패했습니다: {e}") from e

    # ── 2. 출력 파일 열기 ──────────────────────────────────────────────────
    log_callback("info", f"📍 '{dst_path}' 엑셀 파일을 엽니다...")
    is_macro = ext.lower() == ".xlsm"
    try:
        wb = load_workbook(dst_path, keep_vba=is_macro)
    except Exception as e:
        raise RuntimeError(f"Excel 파일을 열 수 없습니다: {e}") from e

    missing = [s for s in SHEETS if s not in wb.sheetnames]
    if missing:
        wb.close()
        raise ValueError(
            f"다음 시트를 찾을 수 없습니다: {', '.join(missing)} "
            f"| 현재 시트: {', '.join(wb.sheetnames)}"
        )

    sheets = [wb[s] for s in SHEETS]

    try:
        # ── 3. A~H 공통정보 기록 ───────────────────────────────────────────
        log_callback("info", f"📍 총 {len(rows_list)}건 A~H 공통정보 기록 중...")
        _write_common_info(sheets, rows_list)

        # ── 4. L/M/N 오류항목 분배 ────────────────────────────────────────
        log_callback("info", f"📍 총 {len(parsed_items_list)}건 오류항목 분배 중...")
        for offset, raw_items in enumerate(parsed_items_list):
            _distribute_and_write(raw_items, DATA_START_ROW + offset, sheets)

        # ── 5. 저장 ────────────────────────────────────────────────────────
        log_callback("info", "📍 저장 중...")
        try:
            wb.save(dst_path)
        except Exception as e:
            raise RuntimeError(f"파일 저장에 실패했습니다: {e}") from e

    finally:
        wb.close()

    log_callback("info", f"✅ 오류항목 분배 완료. → '{dst_path}'")
    return dst_path


def process_distribution(src_path: str, log_callback: Callable = None) -> str:
    """(독립 실행용) 1차 점검 시트의 L/M/N을 읽어 2~4차 시트로 분배합니다.

    pipeline과 무관하게 단독으로 사용 가능합니다.
    A~K열은 이미 기록되어 있다고 가정하고 L/M/N 분배만 수행합니다.

    입력: .xlsx 또는 .xlsm
    출력: 입력과 동일한 확장자로 _자동분류 접미사 추가
          예) 파일.xlsm → 파일_자동분류.xlsm
    """
    log_callback = log_callback or default_log_callback

    if not src_path:
        raise ValueError("파일 경로가 비어 있습니다.")
    if not os.path.exists(src_path):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {src_path}")
    _, ext = os.path.splitext(src_path)
    if ext.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(f".xlsx 또는 .xlsm 파일만 지원됩니다: {src_path}")

    base     = os.path.splitext(src_path)[0]
    dst_path = f"{base}{OUTPUT_SUFFIX}{ext}"

    log_callback("info", "📍 파일 로딩 중...")
    is_macro = ext.lower() == ".xlsm"
    try:
        wb = load_workbook(src_path, keep_vba=is_macro)
    except Exception as e:
        raise RuntimeError(f"Excel 파일을 열 수 없습니다: {e}") from e

    missing = [s for s in SHEETS if s not in wb.sheetnames]
    if missing:
        wb.close()
        raise ValueError(
            f"다음 시트를 찾을 수 없습니다: {', '.join(missing)} "
            f"| 현재 시트: {', '.join(wb.sheetnames)}"
        )

    ws1, ws2, ws3, ws4 = [wb[s] for s in SHEETS]

    log_callback("info", "📍 데이터 범위 분석 중...")
    last_row = DATA_START_ROW - 1
    for r in range(DATA_START_ROW, ws1.max_row + 1):
        if any(ws1.cell(r, c).value not in (None, "") for c in (_COL_ERR, _COL_PAGE, _COL_FIX)):
            last_row = r

    try:
        if last_row < DATA_START_ROW:
            log_callback("info", "ℹ️ 처리할 오류항목이 없습니다. 그대로 저장합니다.")
            wb.save(dst_path)
            return dst_path

        total_rows = last_row - DATA_START_ROW + 1
        log_callback("info", f"📍 총 {total_rows}행 오류항목 분배 중...")

        sheets = [ws1, ws2, ws3, ws4]
        for r in range(DATA_START_ROW, last_row + 1):
            raw_items = _row_cells_to_raw_items(ws1, r)
            _distribute_and_write(raw_items, r, sheets)

        log_callback("info", "📍 저장 중...")
        try:
            wb.save(dst_path)
        except Exception as e:
            raise RuntimeError(f"파일 저장에 실패했습니다: {e}") from e

    finally:
        wb.close()

    log_callback("info", f"✅ 저장 완료. → '{dst_path}'")
    return dst_path
