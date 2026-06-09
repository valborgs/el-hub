# -*- coding: utf-8 -*-
"""
scrape_distr_pipeline/design_core.py  ── 3단계: 서식 일괄 적용

엑셀 파일의 4개 시트 데이터 영역(A~O열)에 테두리·폰트·정렬·행 높이를
일괄 적용합니다.

공개 함수
    run_design(excel_file, log_callback)
        → 엑셀 파일을 열어 서식 적용 후 동일 경로에 저장합니다.
          다른 모듈에 대한 의존성 없이 단독 실행 가능합니다.

열별 정렬 규칙
    A~D (1~4)   : 줄바꿈 없음, 가운데 정렬
    E~K (5~11)  : 줄바꿈 없음, 일반 정렬
    L~M (12~13) : 줄바꿈, 가운데 정렬
    N~O (14~15) : 줄바꿈, 일반 정렬
"""

import os
from typing import Callable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from common.constants import DATA_START_ROW, SHEETS
from common.utils import default_log_callback

# ---------------------------------------------------------------------------
# 서식 상수
# ---------------------------------------------------------------------------
_DATA_FONT = Font(name="맑은 고딕", size=10)
_ALIGN_NOWRAP_CENTER = Alignment(wrap_text=False, vertical="center", horizontal="center")
_ALIGN_NOWRAP_GENERAL = Alignment(wrap_text=False, vertical="center")
_ALIGN_WRAP_CENTER = Alignment(wrap_text=True,  vertical="top",    horizontal="center")
_ALIGN_WRAP_GENERAL = Alignment(wrap_text=True,  vertical="top")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ---------------------------------------------------------------------------
# 내부 함수
# ---------------------------------------------------------------------------

def _find_last_row(wb) -> int:
    """4개 시트 중 A열 기준 데이터가 있는 가장 마지막 행 번호를 반환합니다."""
    last_row = DATA_START_ROW - 1
    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row_num in range(ws.max_row, DATA_START_ROW - 1, -1):
            if ws.cell(row_num, 1).value not in (None, ""):
                if row_num > last_row:
                    last_row = row_num
                break
    return last_row


def _alignment_for_column(col_num: int) -> Alignment:
    """컬럼 번호에 해당하는 Alignment 객체를 반환합니다."""
    if col_num in (1, 2, 3, 4):
        return _ALIGN_NOWRAP_CENTER
    if col_num in (12, 13):
        return _ALIGN_WRAP_CENTER
    if col_num <= 11:
        return _ALIGN_NOWRAP_GENERAL
    return _ALIGN_WRAP_GENERAL


def _apply_format(wb, last_row: int) -> None:
    """DATA_START_ROW~last_row 범위의 A~O열에 서식을 일괄 적용합니다."""
    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row_num in range(DATA_START_ROW, last_row + 1):
            ws.row_dimensions[row_num].auto_size = True
            for col_num in range(1, 16):    # A ~ O (15열)
                cell = ws.cell(row_num, col_num)
                cell.border    = _THIN_BORDER
                cell.font      = _DATA_FONT
                cell.alignment = _alignment_for_column(col_num)


# ---------------------------------------------------------------------------
# 공개 API
# ---------------------------------------------------------------------------

def run_design(excel_file: str, log_callback: Callable = None) -> str:
    """엑셀 파일 4개 시트의 데이터 영역에 서식을 일괄 적용하고 저장합니다.

    Args:
        excel_file:   서식을 적용할 엑셀 파일 경로 (.xlsx 또는 .xlsm)
        log_callback: log_callback(status: str, message: str) 패턴

    Returns:
        서식이 적용된 파일 경로 (입력 경로와 동일)

    Raises:
        FileNotFoundError: 파일이 존재하지 않을 때
        RuntimeError:      파일 열기·저장 실패 시
    """
    log_callback = log_callback or default_log_callback

    if not excel_file:
        raise ValueError("파일 경로가 비어 있습니다.")
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"파일을 찾을 수 없습니다: {excel_file}")
    _, ext = os.path.splitext(excel_file)
    if ext.lower() not in (".xlsx", ".xlsm"):
        raise ValueError(f".xlsx 또는 .xlsm 파일만 지원됩니다: {excel_file}")
    is_macro = ext.lower() == ".xlsm"

    log_callback("info", f"📍 '{excel_file}' 열기 중...")
    try:
        wb = load_workbook(excel_file, keep_vba=is_macro, rich_text=True)
    except Exception as e:
        raise RuntimeError(f"Excel 파일을 열 수 없습니다: {e}") from e

    try:
        last_row = _find_last_row(wb)
        if last_row < DATA_START_ROW:
            log_callback("info", "ℹ️ 서식을 적용할 데이터가 없습니다.")
            return excel_file

        log_callback("info", f"📍 {last_row - DATA_START_ROW + 1}행 서식(테두리·폰트·정렬·행 높이) 적용 중...")
        _apply_format(wb, last_row)

        log_callback("info", f"📍 '{excel_file}' 저장 중...")
        try:
            wb.save(excel_file)
        except Exception as e:
            raise RuntimeError(f"파일 저장에 실패했습니다: {e}") from e
    finally:
        wb.close()

    log_callback("info", f"✅ 서식 적용 완료. → '{excel_file}'")
    return excel_file
